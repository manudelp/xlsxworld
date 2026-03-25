from __future__ import annotations
from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from pydantic import BaseModel
from typing import List, Any, Dict, Tuple
from openpyxl import load_workbook
from io import BytesIO, StringIO
import secrets
import time

# Simple in-memory workbook cache; replace with redis or disk later
# token -> (timestamp, workbook_bytes, sheet_totals)
_WORKBOOK_STORE: Dict[str, Tuple[float, bytes, Dict[str, int]]] = {}
_MAX_STORE = 32  # naive cap
_TTL_SECONDS = 60 * 15  # 15 minutes
_MAX_UPLOAD_SIZE_BYTES = 20 * 1024 * 1024  # 20 MB

router = APIRouter(prefix="/api/inspect", tags=["inspect"])

class SheetPreview(BaseModel):
    name: str
    headers: List[Any]
    sample: List[List[Any]]  # sample rows (excluding header)
    total_rows: int  # including header

class WorkbookPreview(BaseModel):
    token: str
    sheets: List[SheetPreview]
    sheet_count: int

class SheetPage(BaseModel):
    sheet: str
    header: List[Any] | None = None
    rows: List[List[Any]]
    offset: int
    limit: int
    total_rows: int
    done: bool

@router.post("/preview", response_model=WorkbookPreview)
async def preview_workbook(
    file: UploadFile = File(..., description="XLSX file to inspect"),
    sample_rows: int = Query(25, ge=1, le=500),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="Unsupported file type")

    raw = await file.read()

    if len(raw) > _MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"File too large (max {_MAX_UPLOAD_SIZE_BYTES // 1024 // 1024}MB)",
        )

    try:
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {e}")

    # store raw bytes with token
    token = secrets.token_urlsafe(16)
    if len(_WORKBOOK_STORE) >= _MAX_STORE:
        # drop oldest
        oldest = sorted(_WORKBOOK_STORE.items(), key=lambda kv: kv[1][0])[:4]
        for k, _ in oldest:
            _WORKBOOK_STORE.pop(k, None)

    # Build a sheet->row count map (including header row if present)
    sheet_totals: Dict[str, int] = {}

    sheets: List[SheetPreview] = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
            header_list = list(header) if header else []
        except StopIteration:
            header_list = []
            rows_iter = iter(())

        data_rows = []
        total_data_rows = 0
        for r in rows_iter:
            total_data_rows += 1
            if len(data_rows) < sample_rows:
                data_rows.append([*r])

        sheet_total = total_data_rows + (1 if header_list else 0)
        sheet_totals[ws.title] = sheet_total

        sheets.append(
            SheetPreview(
                name=ws.title,
                headers=[*header_list],
                sample=data_rows,
                total_rows=sheet_total,
            )
        )

    _WORKBOOK_STORE[token] = (time.time(), raw, sheet_totals)
    return WorkbookPreview(token=token, sheets=sheets, sheet_count=len(sheets))


def _load_workbook_from_token(token: str):
    meta = _WORKBOOK_STORE.get(token)
    if not meta:
        raise HTTPException(status_code=404, detail="Unknown or expired token")

    if len(meta) == 3:
        ts, raw, sheet_totals = meta
    else:
        ts, raw = meta  # backward compatibility
        sheet_totals = None

    if time.time() - ts > _TTL_SECONDS:
        _WORKBOOK_STORE.pop(token, None)
        raise HTTPException(status_code=404, detail="Token expired")

    # touch
    _WORKBOOK_STORE[token] = (time.time(), raw, sheet_totals) if sheet_totals is not None else (time.time(), raw)

    try:
        return load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to reopen workbook: {e}")


def _get_sheet_total_rows(token: str, sheet: str) -> int | None:
    meta = _WORKBOOK_STORE.get(token)
    if not meta:
        return None

    if len(meta) == 3:
        _, _, sheet_totals = meta
        return sheet_totals.get(sheet)

    return None


@router.get("/sheet", response_model=SheetPage)
async def page_sheet(
    token: str = Query(..., description="Workbook token from preview"),
    sheet: str = Query(..., description="Sheet name"),
    offset: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=2000),
):
    wb = _load_workbook_from_token(token)
    if sheet not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws = wb[sheet]

    sheet_total = _get_sheet_total_rows(token, sheet)
    total_rows = sheet_total if sheet_total is not None else 0
    known_data_rows = sheet_total - 1 if sheet_total is not None and sheet_total > 0 else None

    rows_iter = ws.iter_rows(values_only=True)
    header = None
    data_rows: List[List[Any]] = []

    try:
        first = next(rows_iter)
        header = list(first) if first else []
        if sheet_total is None:
            total_rows += 1
    except StopIteration:
        return SheetPage(sheet=sheet, header=None, rows=[], offset=offset, limit=limit, total_rows=0, done=True)

    # Skip until offset (page starts after header)
    skipped = 0
    while skipped < offset:
        try:
            next(rows_iter)
            skipped += 1
            if sheet_total is None:
                total_rows += 1
        except StopIteration:
            break

    fetched = 0
    for r in rows_iter:
        if fetched < limit:
            data_rows.append([*r])
            fetched += 1
            if sheet_total is None:
                total_rows += 1
        else:
            if sheet_total is None:
                total_rows += 1
                continue
            # known total, we can stop early
            break

    if known_data_rows is not None:
        done = offset + fetched >= known_data_rows
    else:
        # total_rows includes header row if exists
        data_total = max(0, total_rows - 1)
        done = offset + fetched >= data_total

    return SheetPage(
        sheet=sheet,
        header=header,
        rows=data_rows,
        offset=offset,
        limit=limit,
        total_rows=total_rows,
        done=done,
    )

@router.get("/export/csv")
async def export_csv(token: str, sheet: str):
    from fastapi.responses import StreamingResponse
    import csv

    wb = _load_workbook_from_token(token)
    if sheet not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws = wb[sheet]

    def gen():
        # csv.writer writes text, so use a text buffer and encode to bytes per chunk
        output = StringIO()
        writer = csv.writer(output)
        for r in ws.iter_rows(values_only=True):
            writer.writerow(["" if c is None else c for c in r])
            data = output.getvalue()
            if data:
                yield data.encode("utf-8")
            output.seek(0)
            output.truncate(0)

    return StreamingResponse(
        gen(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={sheet}.csv"},
    )

@router.get("/export/json")
async def export_json(token: str, sheet: str):
    from fastapi.responses import StreamingResponse
    import json

    wb = _load_workbook_from_token(token)
    if sheet not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws = wb[sheet]

    def gen():
        first = True
        yield "["
        for r in ws.iter_rows(values_only=True):
            row = ["" if c is None else c for c in r]
            if not first:
                yield ","
            yield json.dumps(row, ensure_ascii=False)
            first = False
        yield "]"

    return StreamingResponse(
        gen(),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={sheet}.json"},
    )

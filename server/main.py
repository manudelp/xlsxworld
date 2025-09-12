from typing import Union

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from io import BytesIO
from . import schemas  # type: ignore

app = FastAPI(title="ilovexlsx API", version="0.1.0")

# TODO: tighten origins for production
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # replace with ["http://localhost:3000"] when known
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/", tags=["root"])
def read_root():
    return {"message": "ilovexlsx backend running"}


@app.get("/health", tags=["meta"])
def health():
    return {"status": "ok"}


@app.get("/items/{item_id}", tags=["example"])
def read_item(item_id: int, q: Union[str, None] = None):
    return {"item_id": item_id, "q": q}


@app.post("/api/inspect/preview", response_model=schemas.WorkbookPreview, tags=["inspect"])
async def preview_workbook(
    file: UploadFile = File(..., description="XLSX file to inspect"),
    sample_rows: int = 25,
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="Unsupported file type")

    raw = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # pragma: no cover - broad catch for parse errors
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {e}")

    sheets: list[schemas.SheetPreview] = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
            header_list = list(header) if header else []
        except StopIteration:
            header_list = []
            rows_iter = iter(())

        data_rows = []
        total_data_rows = 0
        for r in rows_iter:
            total_data_rows += 1
            if len(data_rows) < sample_rows:
                data_rows.append([*r])

        sheets.append(
            schemas.SheetPreview(
                name=ws.title,
                headers=[*header_list],
                sample=data_rows,
                total_rows=total_data_rows + (1 if header_list else 0),
            )
        )

    return schemas.WorkbookPreview(sheets=sheets, sheet_count=len(sheets))
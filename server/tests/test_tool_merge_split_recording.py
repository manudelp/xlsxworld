"""End-to-end recording-wiring coverage for ``merge/*`` + ``split/*``.

Each parametrized case asserts the same contract as the trim-spaces
test in Task 5 (anonymous passthrough + authenticated recording):

* ``POST /append-workbooks`` — multi-file input; recording uses the
  first uploaded filename as ``original_filename``.
* ``POST /merge-sheets``
* ``POST /split-sheet``
* ``POST /split-workbook`` — ZIP output (``application/zip``).
"""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field
from typing import Any, Callable
from unittest.mock import AsyncMock

import openpyxl
import pytest
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient

from app.core.security import AuthenticatedPrincipal, get_current_user_optional
from app.tools._recording import jobs_service_dep
from app.tools.merge.append_workbooks import router as append_workbooks_router
from app.tools.merge.merge_sheets import router as merge_sheets_router
from app.tools.split.split_sheet import router as split_sheet_router
from app.tools.split.split_workbook import router as split_workbook_router


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_two_sheets() -> bytes:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["name", "city"])
    ws1.append(["Alice", "Madrid"])
    ws1.append(["Bob", "Paris"])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["name", "city"])
    ws2.append(["Carol", "Rome"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_single_sheet(rows: int = 6) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", "score"])
    for i in range(1, rows + 1):
        ws.append([f"user_{i}", i * 10])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass
class _Case:
    path: str
    router_factory: Callable[[], Any]
    tool_slug: str
    tool_name: str
    expected_mime: str
    expected_filename: str
    # Callable returning the ``files`` dict for httpx multipart upload.
    files_factory: Callable[[], Any]
    # The filename we expect record_and_respond to see as ``original_filename``.
    expected_original_filename: str
    form_data: dict[str, str] = field(default_factory=dict)


def _merge_sheets_files() -> dict[str, tuple[str, bytes, str]]:
    return {"file": ("workbook.xlsx", _xlsx_two_sheets(), XLSX_MIME)}


def _split_sheet_files() -> dict[str, tuple[str, bytes, str]]:
    return {"file": ("input.xlsx", _xlsx_single_sheet(rows=6), XLSX_MIME)}


def _split_workbook_files() -> dict[str, tuple[str, bytes, str]]:
    return {"file": ("input.xlsx", _xlsx_two_sheets(), XLSX_MIME)}


def _append_workbooks_files() -> list[tuple[str, tuple[str, bytes, str]]]:
    """httpx accepts a list of ``(field_name, value)`` tuples for repeated form keys."""
    payload = _xlsx_two_sheets()
    return [
        ("files", ("first.xlsx", payload, XLSX_MIME)),
        ("files", ("second.xlsx", payload, XLSX_MIME)),
    ]


_CASES: list[_Case] = [
    _Case(
        path="/append-workbooks",
        router_factory=lambda: append_workbooks_router,
        tool_slug="append-workbooks",
        tool_name="Append Workbooks",
        expected_mime=XLSX_MIME,
        expected_filename="appended.xlsx",
        files_factory=_append_workbooks_files,
        expected_original_filename="first.xlsx",
    ),
    _Case(
        path="/merge-sheets",
        router_factory=lambda: merge_sheets_router,
        tool_slug="merge-sheets",
        tool_name="Merge Sheets",
        expected_mime=XLSX_MIME,
        expected_filename="merged.xlsx",
        files_factory=_merge_sheets_files,
        expected_original_filename="workbook.xlsx",
        form_data={"sheet_names": "Sheet1,Sheet2", "output_sheet": "Merged"},
    ),
    _Case(
        path="/split-sheet",
        router_factory=lambda: split_sheet_router,
        tool_slug="split-sheet",
        tool_name="Split Sheet",
        expected_mime=XLSX_MIME,
        expected_filename="split.xlsx",
        files_factory=_split_sheet_files,
        expected_original_filename="input.xlsx",
        form_data={
            "sheet": "Sheet1",
            "chunk_size": "3",
            "part_base": "part",
            "part_separator": "_",
            "numbering_style": "numeric",
        },
    ),
    _Case(
        path="/split-workbook",
        router_factory=lambda: split_workbook_router,
        tool_slug="split-workbook",
        tool_name="Split Workbook",
        expected_mime="application/zip",
        expected_filename="split_workbook.zip",
        files_factory=_split_workbook_files,
        expected_original_filename="input.xlsx",
    ),
]


def _build_app(case: _Case) -> FastAPI:
    app = FastAPI()
    app.include_router(case.router_factory())
    return app


@pytest.mark.asyncio
@pytest.mark.parametrize("case", _CASES, ids=[c.tool_slug for c in _CASES])
async def test_merge_split_anonymous_does_not_record(case: _Case) -> None:
    app = _build_app(case)
    recorded = AsyncMock()
    app.dependency_overrides[get_current_user_optional] = lambda: None
    app.dependency_overrides[jobs_service_dep] = lambda: recorded

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        response = await client.post(
            case.path,
            files=case.files_factory(),
            data=case.form_data,
        )

    assert response.status_code == 200, response.text
    assert "X-Job-Id" not in response.headers
    recorded.record_authenticated_job.assert_not_called()


@pytest.mark.asyncio
@pytest.mark.parametrize("case", _CASES, ids=[c.tool_slug for c in _CASES])
async def test_merge_split_authenticated_schedules_recording(case: _Case) -> None:
    app = _build_app(case)
    principal = AuthenticatedPrincipal(
        user_id=uuid.uuid4(),
        email="user@example.com",
        role="authenticated",
        session_id=None,
        claims={"sub": "user"},
    )
    service = AsyncMock()
    service.record_authenticated_job = AsyncMock(return_value=uuid.uuid4())

    app.dependency_overrides[get_current_user_optional] = lambda: principal
    app.dependency_overrides[jobs_service_dep] = lambda: service

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        response = await client.post(
            case.path,
            files=case.files_factory(),
            data=case.form_data,
            headers={"Authorization": "Bearer any-token"},
        )

    assert response.status_code == 200, response.text
    service.record_authenticated_job.assert_awaited_once()

    kwargs = service.record_authenticated_job.call_args.kwargs
    assert kwargs["user_id"] == principal.user_id
    assert kwargs["tool_slug"] == case.tool_slug
    assert kwargs["tool_name"] == case.tool_name
    assert kwargs["original_filename"] == case.expected_original_filename
    assert kwargs["output_filename"] == case.expected_filename
    assert kwargs["mime_type"] == case.expected_mime
    assert kwargs["success"] is True
    assert kwargs["error_type"] is None
    assert isinstance(kwargs["duration_ms"], int) and kwargs["duration_ms"] >= 0
    assert isinstance(kwargs["output_bytes"], bytes) and len(kwargs["output_bytes"]) > 0

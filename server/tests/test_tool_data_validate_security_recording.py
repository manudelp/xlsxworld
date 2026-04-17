"""End-to-end recording-wiring coverage for ``data/*``, ``validate/*``
and ``security/*`` (batch 5).

Covers seven endpoints:

* ``POST /sort-rows``
* ``POST /split-column``
* ``POST /transpose-sheet``
* ``POST /detect-blanks``     (dynamic ``blanks-report-<base>.xlsx``)
* ``POST /validate-emails``   (dynamic ``email-validation-<base>.xlsx``)
* ``POST /password-protect``
* ``POST /remove-password``

Each tool gets one anonymous case (no recording) and one authenticated
case (recording scheduled with the right slug/name/filenames).
"""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field
from typing import Any, Callable
from unittest.mock import AsyncMock

import openpyxl
import pytest
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient

from app.core.security import AuthenticatedPrincipal, get_current_user_optional
from app.tools._recording import jobs_service_dep
from app.tools.data.sort_rows import router as sort_rows_router
from app.tools.data.split_column import router as split_column_router
from app.tools.data.transpose_sheet import router as transpose_sheet_router
from app.tools.security.password_protect import router as password_protect_router
from app.tools.security.remove_password import router as remove_password_router
from app.tools.validate.detect_blanks import router as detect_blanks_router
from app.tools.validate.validate_emails import router as validate_emails_router


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_sortable() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Score"])
    ws.append(["Bob", 10])
    ws.append(["Alice", 30])
    ws.append(["Carol", 20])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_split_column() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Tags"])
    ws.append(["Alice", "a,b,c"])
    ws.append(["Bob", "x,y"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_with_blanks() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "City"])
    ws.append(["Alice", "Madrid"])
    ws.append(["Bob", None])
    ws.append([None, "Rome"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_with_emails() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Email"])
    ws.append(["Alice", "alice@example.com"])
    ws.append(["Bob", "not-an-email"])
    ws.append(["Carol", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_plain() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Score"])
    ws.append(["Alice", 10])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_protected() -> bytes:
    """Workbook with sheet-level protection enabled (needed by
    ``remove-password``: openpyxl's ``SheetProtection.password``
    setter hashes the incoming value and fails on unprotected sheets
    whose ``password`` slot is ``None``).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Score"])
    ws.append(["Alice", 10])
    ws.protection.sheet = True
    ws.protection.password = "Secret-1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass
class _Case:
    path: str
    router_factory: Callable[[], Any]
    tool_slug: str
    tool_name: str
    files_factory: Callable[[], Any]
    expected_original_filename: str
    expected_mime: str = XLSX_MIME
    expected_filename: str | None = None
    expected_filename_prefix: str | None = None
    form_data: dict[str, str] = field(default_factory=dict)


_CASES: list[_Case] = [
    _Case(
        path="/sort-rows",
        router_factory=lambda: sort_rows_router,
        tool_slug="sort-rows",
        tool_name="Sort Rows",
        files_factory=lambda: {"file": ("input.xlsx", _xlsx_sortable(), XLSX_MIME)},
        expected_original_filename="input.xlsx",
        expected_filename="sorted.xlsx",
        form_data={
            "sheet": "Sheet1",
            "sort_keys": '[{"column":"Name","direction":"asc"}]',
            "has_header": "true",
        },
    ),
    _Case(
        path="/split-column",
        router_factory=lambda: split_column_router,
        tool_slug="split-column",
        tool_name="Split Column",
        files_factory=lambda: {"file": ("input.xlsx", _xlsx_split_column(), XLSX_MIME)},
        expected_original_filename="input.xlsx",
        expected_filename="split-column.xlsx",
        form_data={
            "sheet": "Sheet1",
            "column": "Tags",
            "delimiter": "comma",
            "keep_original": "true",
        },
    ),
    _Case(
        path="/transpose-sheet",
        router_factory=lambda: transpose_sheet_router,
        tool_slug="transpose-sheet",
        tool_name="Transpose Sheet",
        files_factory=lambda: {"file": ("input.xlsx", _xlsx_plain(), XLSX_MIME)},
        expected_original_filename="input.xlsx",
        expected_filename="transposed.xlsx",
        form_data={"sheet": "Sheet1"},
    ),
    _Case(
        path="/detect-blanks",
        router_factory=lambda: detect_blanks_router,
        tool_slug="detect-blanks",
        tool_name="Detect Blanks",
        files_factory=lambda: {"file": ("input.xlsx", _xlsx_with_blanks(), XLSX_MIME)},
        expected_original_filename="input.xlsx",
        expected_filename_prefix="blanks-report-",
    ),
    _Case(
        path="/validate-emails",
        router_factory=lambda: validate_emails_router,
        tool_slug="validate-emails",
        tool_name="Validate Emails",
        files_factory=lambda: {"file": ("input.xlsx", _xlsx_with_emails(), XLSX_MIME)},
        expected_original_filename="input.xlsx",
        expected_filename_prefix="email-validation-",
        form_data={"sheet": "Sheet1", "column": "Email"},
    ),
    _Case(
        path="/password-protect",
        router_factory=lambda: password_protect_router,
        tool_slug="password-protect",
        tool_name="Password Protect",
        files_factory=lambda: {"file": ("input.xlsx", _xlsx_plain(), XLSX_MIME)},
        expected_original_filename="input.xlsx",
        expected_filename="protected.xlsx",
        form_data={"password": "Secret-1"},
    ),
    _Case(
        path="/remove-password",
        router_factory=lambda: remove_password_router,
        tool_slug="remove-password",
        tool_name="Remove Password",
        files_factory=lambda: {"file": ("input.xlsx", _xlsx_protected(), XLSX_MIME)},
        expected_original_filename="input.xlsx",
        expected_filename="unprotected.xlsx",
    ),
]


def _build_app(case: _Case) -> FastAPI:
    app = FastAPI()
    app.include_router(case.router_factory())
    return app


@pytest.mark.asyncio
@pytest.mark.parametrize("case", _CASES, ids=[c.tool_slug for c in _CASES])
async def test_data_validate_security_anonymous_does_not_record(case: _Case) -> None:
    app = _build_app(case)
    recorded = AsyncMock()
    app.dependency_overrides[get_current_user_optional] = lambda: None
    app.dependency_overrides[jobs_service_dep] = lambda: recorded

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        response = await client.post(
            case.path,
            files=case.files_factory(),
            data=case.form_data,
        )

    assert response.status_code == 200, response.text
    assert "X-Job-Id" not in response.headers
    recorded.record_authenticated_job.assert_not_called()


@pytest.mark.asyncio
@pytest.mark.parametrize("case", _CASES, ids=[c.tool_slug for c in _CASES])
async def test_data_validate_security_authenticated_schedules_recording(case: _Case) -> None:
    app = _build_app(case)
    principal = AuthenticatedPrincipal(
        user_id=uuid.uuid4(),
        email="user@example.com",
        role="authenticated",
        session_id=None,
        claims={"sub": "user"},
    )
    service = AsyncMock()
    service.record_authenticated_job = AsyncMock(return_value=uuid.uuid4())

    app.dependency_overrides[get_current_user_optional] = lambda: principal
    app.dependency_overrides[jobs_service_dep] = lambda: service

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        response = await client.post(
            case.path,
            files=case.files_factory(),
            data=case.form_data,
            headers={"Authorization": "Bearer any-token"},
        )

    assert response.status_code == 200, response.text
    service.record_authenticated_job.assert_awaited_once()

    kwargs = service.record_authenticated_job.call_args.kwargs
    assert kwargs["user_id"] == principal.user_id
    assert kwargs["tool_slug"] == case.tool_slug
    assert kwargs["tool_name"] == case.tool_name
    assert kwargs["original_filename"] == case.expected_original_filename
    if case.expected_filename is not None:
        assert kwargs["output_filename"] == case.expected_filename
    if case.expected_filename_prefix is not None:
        assert kwargs["output_filename"].startswith(case.expected_filename_prefix), (
            kwargs["output_filename"]
        )
        assert kwargs["output_filename"].endswith(".xlsx"), kwargs["output_filename"]
    assert kwargs["mime_type"] == case.expected_mime
    assert kwargs["success"] is True
    assert kwargs["error_type"] is None
    assert isinstance(kwargs["duration_ms"], int) and kwargs["duration_ms"] >= 0
    assert isinstance(kwargs["output_bytes"], bytes) and len(kwargs["output_bytes"]) > 0

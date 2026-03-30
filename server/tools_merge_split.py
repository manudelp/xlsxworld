from __future__ import annotations
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO
import re
import zipfile
from openpyxl import load_workbook, Workbook

router = APIRouter(prefix="/api/tools", tags=["tools"])
_MAX_UPLOAD_SIZE_BYTES = 20 * 1024 * 1024  # 20 MB


def _check_file(file: UploadFile, expected_ext: str):
    if not file.filename.lower().endswith(expected_ext):
        raise HTTPException(status_code=400, detail=f"Unsupported file type, expected {expected_ext}")


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _safe_sheet_title(raw: str | None, fallback: str) -> str:
    value = (raw or "").strip()
    if not value:
        value = fallback

    # Excel sheet title rules: max 31 chars, no \\ / * ? : [ ]
    value = _INVALID_SHEET_CHARS.sub("_", value).strip("'")
    if not value:
        value = fallback

    return value[:31]


def _unique_sheet_title(base: str, used: set[str]) -> str:
    if base not in used:
        used.add(base)
        return base

    for index in range(2, 10_000):
        suffix = f"_{index}"
        allowed = max(1, 31 - len(suffix))
        candidate = f"{base[:allowed]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate

    # extremely defensive fallback
    candidate = f"part_{len(used) + 1}"[:31]
    used.add(candidate)
    return candidate


def _alpha_token(index: int, lowercase: bool = False) -> str:
    # 1 -> A, 26 -> Z, 27 -> AA
    chars: list[str] = []
    n = max(1, index)
    while n > 0:
        n -= 1
        chars.append(chr(ord("A") + (n % 26)))
        n //= 26
    token = "".join(reversed(chars))
    return token.lower() if lowercase else token


def _roman_token(index: int, lowercase: bool = False) -> str:
    n = max(1, index)
    numerals = [
        (1000, "M"),
        (900, "CM"),
        (500, "D"),
        (400, "CD"),
        (100, "C"),
        (90, "XC"),
        (50, "L"),
        (40, "XL"),
        (10, "X"),
        (9, "IX"),
        (5, "V"),
        (4, "IV"),
        (1, "I"),
    ]
    parts: list[str] = []
    for value, symbol in numerals:
        while n >= value:
            parts.append(symbol)
            n -= value
    token = "".join(parts)
    return token.lower() if lowercase else token


def _build_part_token(
    index: int,
    numbering_style: str,
    custom_tokens: list[str],
) -> str:
    if numbering_style == "numeric":
        return str(index)
    if numbering_style == "numeric-padded":
        return f"{index:02d}"
    if numbering_style == "alpha-upper":
        return _alpha_token(index, lowercase=False)
    if numbering_style == "alpha-lower":
        return _alpha_token(index, lowercase=True)
    if numbering_style == "roman-upper":
        return _roman_token(index, lowercase=False)
    if numbering_style == "roman-lower":
        return _roman_token(index, lowercase=True)
    if numbering_style == "custom":
        if index <= len(custom_tokens):
            return custom_tokens[index - 1]
        return str(index)
    return str(index)


@router.post("/merge-sheets")
async def merge_sheets(
    file: UploadFile = File(..., description="XLSX file"),
    sheet_names: str = Form("", description="Comma-separated sheet names to merge (empty=all)"),
    output_sheet: str = Form("Merged", description="Output sheet name"),
):
    _check_file(file, ".xlsx")

    raw = await file.read()
    if len(raw) > _MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(status_code=400, detail="File too large")

    try:
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {e}")

    selected = [s.strip() for s in sheet_names.split(",") if s.strip()]
    if not selected:
        selected = wb.sheetnames

    for s in selected:
        if s not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet not found: {s}")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = output_sheet[:31] if output_sheet else "Merged"

    header_written = False
    row_count = 0
    for sheet_name in selected:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        for i, row in enumerate(rows):
            if i == 0:
                if not header_written:
                    out_ws.append(["" if v is None else v for v in row])
                    header_written = True
                elif row and any(x is not None for x in row):
                    # skip repeating header for following sheets
                    continue
            else:
                out_ws.append(["" if v is None else v for v in row])
                row_count += 1

    output = BytesIO()
    out_wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=merged.xlsx"},
    )


@router.post("/split-sheet")
async def split_sheet(
    file: UploadFile = File(..., description="XLSX file"),
    sheet: str = Form(..., description="Sheet name"),
    chunk_size: int = Form(1000, description="Max rows per chunk (including header)"),
    part_base: str = Form("part", description="Base name used for split sheets"),
    part_separator: str = Form("_", description="Separator between base name and part token"),
    numbering_style: str = Form("numeric", description="Part token style"),
    custom_sequence: str = Form("", description="Custom tokens (one per line) when style is custom"),
):
    _check_file(file, ".xlsx")

    raw = await file.read()
    if len(raw) > _MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(status_code=400, detail="File too large")
    if chunk_size < 2:
        raise HTTPException(status_code=400, detail="chunk_size must be >= 2")

    try:
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {e}")

    if sheet not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found")

    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Sheet is empty")

    valid_styles = {
        "numeric",
        "numeric-padded",
        "alpha-upper",
        "alpha-lower",
        "roman-upper",
        "roman-lower",
        "custom",
    }
    if numbering_style not in valid_styles:
        raise HTTPException(status_code=400, detail="Invalid numbering_style")

    if len(part_separator) > 4:
        raise HTTPException(status_code=400, detail="part_separator must be 4 characters or fewer")

    custom_tokens = [
        token.strip()
        for token in re.split(r"[\r\n,]+", custom_sequence)
        if token.strip()
    ]
    if numbering_style == "custom" and not custom_tokens:
        raise HTTPException(status_code=400, detail="custom_sequence is required when numbering_style is custom")

    # first row as header
    header = rows[0]
    chunks = [rows[i : i + chunk_size] for i in range(1, len(rows), chunk_size)]

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    used_titles: set[str] = set()

    for idx, chunk in enumerate(chunks, start=1):
        token = _build_part_token(idx, numbering_style, custom_tokens)
        default_name = f"part_{idx}"
        requested_name = f"{part_base}{part_separator}{token}"
        base_title = _safe_sheet_title(requested_name, default_name)
        sheet_title = _unique_sheet_title(base_title, used_titles)
        part = out_wb.create_sheet(sheet_title)
        part.append(["" if v is None else v for v in header])
        for row in chunk:
            part.append(["" if v is None else v for v in row])

    if not chunks:
        # no data rows, create one sheet with header only
        token = _build_part_token(1, numbering_style, custom_tokens)
        requested_name = f"{part_base}{part_separator}{token}"
        base_title = _safe_sheet_title(requested_name, "part_1")
        sheet_title = _unique_sheet_title(base_title, used_titles)
        part = out_wb.create_sheet(sheet_title)
        part.append(["" if v is None else v for v in header])

    output = BytesIO()
    out_wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=splitted.xlsx"},
    )


@router.post("/append-workbooks")
async def append_workbooks(
    files: list[UploadFile] = File(..., description="XLSX files to append"),
):
    if len(files) < 2:
        raise HTTPException(status_code=400, detail="At least two workbook files are required")

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    used_titles: set[str] = set()
    copied_sheets = 0

    for file_index, file in enumerate(files, start=1):
        filename = file.filename or f"workbook_{file_index}.xlsx"
        source_name = filename.rsplit(".", 1)[0] or f"workbook_{file_index}"
        source_name = _INVALID_SHEET_CHARS.sub("_", source_name)

        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Unsupported file type, expected .xlsx")

        raw = await file.read()
        if len(raw) > _MAX_UPLOAD_SIZE_BYTES:
            raise HTTPException(status_code=400, detail="One of the files is too large")

        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
        for sheet in wb.worksheets:
            preferred_title = _safe_sheet_title(sheet.title, f"sheet_{copied_sheets + 1}")
            requested_title = _safe_sheet_title(
                f"{source_name}_{preferred_title}",
                preferred_title,
            )
            target_title = _unique_sheet_title(requested_title, used_titles)
            out_ws = out_wb.create_sheet(target_title)

            for row in sheet.iter_rows(values_only=True):
                out_ws.append(["" if v is None else v for v in row])

            copied_sheets += 1

    if copied_sheets == 0:
        raise HTTPException(status_code=400, detail="No data found in uploaded workbooks")

    output = BytesIO()
    out_wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=appended.xlsx"},
    )


@router.post("/split-workbook")
async def split_workbook(file: UploadFile = File(..., description="XLSX file")):
    _check_file(file, ".xlsx")
    raw = await file.read()
    if len(raw) > _MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(status_code=400, detail="File too large")

    wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    if not wb.sheetnames:
        raise HTTPException(status_code=400, detail="Workbook is empty")

    zipped = BytesIO()
    with zipfile.ZipFile(zipped, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            child_wb = Workbook()
            child_ws = child_wb.active
            child_ws.title = sheet_name[:31] if sheet_name else "Sheet1"

            for row in ws.iter_rows(values_only=True):
                child_ws.append(["" if v is None else v for v in row])

            child_bytes = BytesIO()
            child_wb.save(child_bytes)
            child_bytes.seek(0)
            member_name = f"{sheet_name.replace(' ', '_') or 'sheet'}.xlsx"
            zf.writestr(member_name, child_bytes.getvalue())

    zipped.seek(0)
    return StreamingResponse(
        iter([zipped.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=splitted_workbook.zip"},
    )

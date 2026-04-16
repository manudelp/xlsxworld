"""Helpers to edit Excel workbooks while preserving as much formatting as possible.

When a tool only needs to mutate cell *values* (e.g. trim spaces, change case,
find/replace, remove duplicate rows), we should preserve the workbook's
formatting (column widths, row heights, fonts, fills, borders, alignment,
number formats, named ranges, freeze panes, conditional formatting, etc.)
rather than rebuilding the workbook from scratch.

This module provides a small wrapper around openpyxl.load_workbook that:
  - keeps formulas, styles, named ranges, print settings
  - preserves VBA macros for .xlsm files (keep_vba=True)
  - tries hard to keep images/charts; if openpyxl can't parse them, falls back
    to a "stripped" load and reports back so callers can warn the user

For .xls and .xlsb files, in-place editing is not supported by openpyxl, and
callers should fall back to the data-only round-trip (which always works but
loses formatting).
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from app.services.excel_reader import (
    _OPENPYXL_EXTENSIONS,
    _extract_extension,
    _sanitize_openxml_for_data_read,
    _validate_magic_bytes,
    ensure_supported_excel_filename,
)

__all__ = [
    "load_workbook_for_edit",
    "save_workbook_to_bytes",
    "supports_inplace_edit",
    "FormatPreservingLoad",
]


class FormatPreservingLoad:
    """Result of loading a workbook for in-place editing."""

    __slots__ = ("workbook", "visual_elements_lost")

    def __init__(self, workbook: Workbook, *, visual_elements_lost: bool) -> None:
        self.workbook = workbook
        self.visual_elements_lost = visual_elements_lost


def supports_inplace_edit(filename: str | None) -> bool:
    """Return True if this file type can be edited in place by openpyxl."""
    return _extract_extension(filename) in _OPENPYXL_EXTENSIONS


def load_workbook_for_edit(raw: bytes, filename: str | None) -> FormatPreservingLoad:
    """Load an .xlsx/.xlsm/.xltx/.xltm/.xlam workbook for in-place editing.

    Always preserves: cell styles, column dimensions, row dimensions, named
    ranges, print settings, freeze panes, conditional formatting, data
    validations.

    For .xlsm files, VBA macros are preserved (keep_vba=True). Links to
    external workbooks are dropped to avoid network IO during save.

    If openpyxl fails to parse images/charts, we retry against a sanitized
    copy of the workbook that has visual elements stripped, and the returned
    `visual_elements_lost` flag is set so callers can surface a warning.
    """
    extension = _extract_extension(filename)
    ensure_supported_excel_filename(filename)
    _validate_magic_bytes(raw, extension)

    if extension not in _OPENPYXL_EXTENSIONS:
        # Caller should check supports_inplace_edit() first.
        raise HTTPException(
            status_code=400,
            detail="In-place editing is only supported for .xlsx/.xlsm files.",
        )

    keep_vba = extension in {".xlsm", ".xltm", ".xlam"}

    try:
        wb = load_workbook(
            filename=BytesIO(raw),
            data_only=False,
            keep_links=False,
            keep_vba=keep_vba,
        )
        return FormatPreservingLoad(wb, visual_elements_lost=False)
    except Exception:
        # Drawings/charts/images couldn't be parsed. Strip them and retry so
        # the rest of the formatting (column widths, styles, etc.) is kept.
        sanitized = _sanitize_openxml_for_data_read(raw)
        try:
            wb = load_workbook(
                filename=BytesIO(sanitized),
                data_only=False,
                keep_links=False,
                keep_vba=keep_vba,
            )
            return FormatPreservingLoad(wb, visual_elements_lost=True)
        except Exception as error:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Could not read the workbook. The file may be corrupted "
                    "or in an unsupported format."
                ),
            ) from error


def save_workbook_to_bytes(workbook: Workbook) -> bytes:
    """Serialize an openpyxl workbook to bytes."""
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def header_index_map(
    sheet: Any,
    *,
    header_row: int = 1,
    selected_columns: list[str] | None,
) -> list[int]:
    """Return the 1-based column indexes for `selected_columns`.

    If `selected_columns` is empty/None, returns indexes for all non-empty
    header cells. Compares column names case-sensitively against trimmed
    header cell values.
    """
    headers: list[str] = []
    for cell in sheet[header_row]:
        value = cell.value
        headers.append("" if value is None else str(value).strip())

    if not selected_columns:
        return [idx + 1 for idx, name in enumerate(headers) if name]

    indexes: list[int] = []
    for name in selected_columns:
        if name in headers:
            indexes.append(headers.index(name) + 1)
    return indexes

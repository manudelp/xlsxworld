from __future__ import annotations

from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Any
import zipfile
import xml.etree.ElementTree as ET

import xlrd
from fastapi import HTTPException
from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb_workbook

SUPPORTED_EXCEL_EXTENSIONS = (
    ".xlsx",
    ".xls",
    ".xlsm",
    ".xlsb",
    ".xltx",
    ".xltm",
    ".xlam",
)

_OPENPYXL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xlam"}
_XLRD_EXTENSIONS = {".xls"}
_PYXLSB_EXTENSIONS = {".xlsb"}
_DRAWING_PREFIXES = (
    "xl/drawings/",
    "xl/charts/",
    "xl/media/",
)


def _extract_extension(filename: str | None) -> str:
    name = (filename or "").strip().lower()
    if "." not in name:
        return ""
    return f".{name.rsplit('.', 1)[-1]}"


def ensure_supported_excel_filename(filename: str | None):
    extension = _extract_extension(filename)
    if extension not in SUPPORTED_EXCEL_EXTENSIONS:
        expected = ", ".join(SUPPORTED_EXCEL_EXTENSIONS)
        raise HTTPException(status_code=400, detail=f"Unsupported file type, expected one of: {expected}")


_ZIP_MAGIC = b"PK\x03\x04"
_OLE_MAGIC = b"\xd0\xcf\x11\xe0"


def _validate_magic_bytes(raw: bytes, extension: str):
    if extension in _OPENPYXL_EXTENSIONS | _PYXLSB_EXTENSIONS:
        if not raw[:4].startswith(_ZIP_MAGIC):
            raise HTTPException(status_code=400, detail="File content does not match expected format")
    elif extension in _XLRD_EXTENSIONS:
        if not raw[:4].startswith(_OLE_MAGIC):
            raise HTTPException(status_code=400, detail="File content does not match expected format")


def _read_openpyxl_values(raw: bytes) -> dict[str, list[list[Any]]]:
    wb_data = load_workbook(
        filename=BytesIO(raw),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    data: dict[str, list[list[Any]]] = {}
    has_none = False
    for sheet in wb_data.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        data[sheet.title] = rows
        if not has_none and any(cell is None for row in rows for cell in row):
            has_none = True
    wb_data.close()

    if not has_none:
        return data

    wb_formula = load_workbook(
        filename=BytesIO(raw),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    for sheet in wb_formula.worksheets:
        if sheet.title not in data:
            continue
        data_rows = data[sheet.title]
        for r, row in enumerate(sheet.iter_rows(values_only=True)):
            if r >= len(data_rows):
                break
            for c, val in enumerate(row):
                if c < len(data_rows[r]) and data_rows[r][c] is None and val is not None:
                    data_rows[r][c] = val
    wb_formula.close()

    return data


def _strip_visual_relationships(xml_bytes: bytes) -> bytes:
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return xml_bytes

    namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    rel_tag = f"{{{namespace}}}Relationship"
    removed = False

    for rel in list(root.findall(rel_tag)):
        target = (rel.attrib.get("Target") or "").replace("\\", "/").lower()
        rel_type = (rel.attrib.get("Type") or "").lower()

        is_visual_target = (
            "/drawings/" in target
            or "/charts/" in target
            or "/media/" in target
            or target.startswith("../drawings/")
            or target.startswith("../charts/")
            or target.startswith("../media/")
        )
        is_visual_type = any(
            marker in rel_type
            for marker in ("/drawing", "/chart", "/image")
        )

        if is_visual_target or is_visual_type:
            root.remove(rel)
            removed = True

    if not removed:
        return xml_bytes

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _sanitize_openxml_for_data_read(raw: bytes) -> bytes:
    source = BytesIO(raw)
    output = BytesIO()

    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            name = item.filename.replace("\\", "/")
            lower_name = name.lower()

            if lower_name.startswith(_DRAWING_PREFIXES):
                continue

            payload = zin.read(item.filename)
            if lower_name.endswith(".rels"):
                payload = _strip_visual_relationships(payload)

            zout.writestr(name, payload)

    return output.getvalue()


def parse_excel_bytes(raw: bytes, filename: str | None) -> dict[str, list[list[Any]]]:
    extension = _extract_extension(filename)
    ensure_supported_excel_filename(filename)
    _validate_magic_bytes(raw, extension)

    try:
        if extension in _OPENPYXL_EXTENSIONS:
            try:
                return _read_openpyxl_values(raw)
            except Exception:
                sanitized = _sanitize_openxml_for_data_read(raw)
                return _read_openpyxl_values(sanitized)

        if extension in _XLRD_EXTENSIONS:
            workbook = xlrd.open_workbook(file_contents=raw)
            data = {}
            for sheet in workbook.sheets():
                rows: list[list[Any]] = []
                for row_index in range(sheet.nrows):
                    rows.append([sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)])
                data[sheet.name] = rows
            return data

        if extension in _PYXLSB_EXTENSIONS:
            data = {}
            with NamedTemporaryFile(delete=True, suffix=extension) as temp:
                temp.write(raw)
                temp.flush()
                with open_xlsb_workbook(temp.name) as workbook:
                    for sheet_name in workbook.sheets:
                        rows: list[list[Any]] = []
                        with workbook.get_sheet(sheet_name) as sheet:
                            for row in sheet.rows():
                                values = [cell.v for cell in row]
                                while values and values[-1] is None:
                                    values.pop()
                                rows.append(values)
                        data[sheet_name] = rows
            return data

    except HTTPException:
        raise
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {error}") from error

    raise HTTPException(status_code=400, detail="Unsupported workbook format")

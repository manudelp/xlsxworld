from __future__ import annotations

from io import BytesIO
from typing import Any, Callable

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.workbook import Workbook as OpenpyxlWorkbook

from app.services.excel_editor import (
    FormatPreservingLoad,
    header_index_map,
    load_workbook_for_edit,
    save_workbook_to_bytes,
    supports_inplace_edit,
)


def parse_columns_arg(columns: str) -> list[str]:
    return [column.strip() for column in columns.split(",") if column.strip()]


def resolve_target_sheets(
    workbook_data: dict[str, list[list[Any]]],
    sheet: str,
    all_sheets: bool,
) -> list[str]:
    if all_sheets:
        return list(workbook_data.keys())

    selected_sheet = sheet.strip()
    if not selected_sheet:
        raise HTTPException(status_code=400, detail="sheet is required when all_sheets is false")
    if selected_sheet not in workbook_data:
        raise HTTPException(status_code=404, detail="Sheet not found")
    return [selected_sheet]


def resolve_column_indexes(
    header: list[Any],
    selected_columns: list[str],
    *,
    allow_missing: bool = False,
) -> list[int]:
    if not selected_columns:
        return list(range(len(header)))

    indexes: list[int] = []
    missing: list[str] = []

    normalized_header = ["" if value is None else str(value).strip() for value in header]

    for column in selected_columns:
        if column in normalized_header:
            indexes.append(normalized_header.index(column))
        else:
            missing.append(column)

    if missing and not allow_missing:
        raise HTTPException(
            status_code=400,
            detail=f"Column not found: {missing[0]}",
        )

    return indexes


def get_cell(row: list[Any], index: int) -> Any:
    if index < len(row):
        return row[index]
    return None


def with_updated_cell(row: list[Any], index: int, value: Any) -> list[Any]:
    out = list(row)
    if index >= len(out):
        out.extend([None] * (index + 1 - len(out)))
    out[index] = value
    return out


def workbook_bytes_from_data(workbook_data: dict[str, list[list[Any]]]) -> bytes:
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for sheet_name, rows in workbook_data.items():
        out_ws = out_wb.create_sheet(sheet_name[:31] if sheet_name else "Sheet")
        for row in rows:
            out_ws.append(["" if cell is None else cell for cell in row])

    if not out_wb.worksheets:
        out_wb.create_sheet("Sheet1")

    output = BytesIO()
    out_wb.save(output)
    return output.getvalue()


def resolve_target_sheet_titles(
    workbook: OpenpyxlWorkbook,
    sheet: str,
    all_sheets: bool,
) -> list[str]:
    if all_sheets:
        return list(workbook.sheetnames)

    selected = sheet.strip()
    if not selected:
        raise HTTPException(
            status_code=400,
            detail="sheet is required when all_sheets is false",
        )
    if selected not in workbook.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found")
    return [selected]


def apply_value_mutation_inplace(
    raw: bytes,
    filename: str,
    *,
    sheet: str,
    all_sheets: bool,
    selected_columns: list[str],
    mutate: Callable[[Any], Any],
) -> tuple[bytes, bool]:
    """Mutate string values in selected cells while preserving formatting.

    The provided `mutate` callback is invoked for every data cell (non-header)
    in the targeted columns. It receives the current cell value and must
    return the new value. Non-string values are passed through unchanged
    unless the callback explicitly handles them.

    Returns (output_bytes, visual_elements_removed).
    """
    if not supports_inplace_edit(filename):
        return None, False  # type: ignore[return-value]

    loaded: FormatPreservingLoad = load_workbook_for_edit(raw, filename)
    workbook = loaded.workbook

    target_sheets = resolve_target_sheet_titles(workbook, sheet, all_sheets)

    for sheet_name in target_sheets:
        ws = workbook[sheet_name]
        if ws.max_row <= 1:
            continue

        column_indexes = header_index_map(
            ws,
            header_row=1,
            selected_columns=selected_columns or None,
        )
        if not column_indexes:
            # If user requested specific columns but none matched on this sheet,
            # skip silently when applying to all sheets; otherwise raise.
            if selected_columns and not all_sheets:
                raise HTTPException(
                    status_code=400,
                    detail=f"Column not found: {selected_columns[0]}",
                )
            continue

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_index in column_indexes:
                # iter_rows is 0-based within the row tuple
                idx0 = col_index - 1
                if idx0 >= len(row):
                    continue
                cell = row[idx0]
                new_value = mutate(cell.value)
                if new_value != cell.value:
                    cell.value = new_value

    output = save_workbook_to_bytes(workbook)
    return output, loaded.visual_elements_lost


def delete_rows_inplace(
    raw: bytes,
    filename: str,
    *,
    sheet: str,
    all_sheets: bool,
    should_delete: Callable[[list[Any]], bool],
) -> tuple[bytes, bool, int]:
    """Delete rows from selected sheets while preserving formatting.

    The `should_delete` callback is invoked once per data row (excluding the
    header at row 1) with the row's values as a Python list, and must return
    True to mark the row for deletion. Returns
    (output_bytes, visual_elements_removed, total_rows_removed).
    """
    loaded: FormatPreservingLoad = load_workbook_for_edit(raw, filename)
    workbook = loaded.workbook

    target_sheets = resolve_target_sheet_titles(workbook, sheet, all_sheets)

    total_removed = 0
    for sheet_name in target_sheets:
        ws = workbook[sheet_name]
        if ws.max_row <= 1:
            continue

        rows_to_delete: list[int] = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            values = [cell.value for cell in row]
            if should_delete(values):
                rows_to_delete.append(row[0].row)

        # Delete from bottom to top so indices stay valid.
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx, 1)
        total_removed += len(rows_to_delete)

    output = save_workbook_to_bytes(workbook)
    return output, loaded.visual_elements_lost, total_removed

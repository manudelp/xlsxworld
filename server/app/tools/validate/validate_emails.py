from __future__ import annotations

import re
import time
from io import BytesIO
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.security import AuthenticatedPrincipal
from app.services.excel_reader import parse_excel_bytes
from app.services.jobs_service import JobsService
from app.tools._common import check_excel_file, read_upload_for_principal, safe_base_filename
from app.tools._recording import (
    get_current_user_optional,
    jobs_service_dep,
    record_and_respond,
)
from app.tools.analyze._styles import (
    ALT_ROW_FILL, BODY_FONT, THIN_BORDER, WHITE_FILL,
    apply_header_row, auto_size,
)

router = APIRouter()

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")

GREEN_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
RED_FILL = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
GRAY_FILL = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)


@router.post(
    "/validate-emails",
    summary="Validate Emails",
    description="Scans selected columns for email addresses and validates format.",
)
async def validate_emails(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="Excel file"),
    sheet: str = Form(..., description="Sheet name"),
    column: str = Form(..., description="Column header name containing emails"),
    principal: AuthenticatedPrincipal | None = Depends(get_current_user_optional),
    jobs_service: JobsService = Depends(jobs_service_dep),
):
    started = time.perf_counter()
    check_excel_file(file)
    raw = await read_upload_for_principal(file, principal=principal)

    workbook_data = parse_excel_bytes(raw, file.filename)

    if sheet not in workbook_data:
        raise HTTPException(status_code=404, detail="Sheet not found")

    rows = workbook_data[sheet]
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Sheet has no data rows")

    header = rows[0]
    data_rows = rows[1:]

    col_idx = None
    for i, h in enumerate(header):
        if str(h).strip() == column.strip():
            col_idx = i
            break
    if col_idx is None:
        raise HTTPException(status_code=400, detail=f"Column '{column}' not found")

    valid_count = 0
    invalid_count = 0
    empty_count = 0

    out = Workbook()
    ws = out.active
    ws.title = "Email Validation"

    out_headers = [str(h) if h is not None else f"Col {i+1}" for i, h in enumerate(header)] + ["Email Status"]
    for c, h in enumerate(out_headers, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_row(ws, 1, len(out_headers))

    status_col = len(out_headers)

    for idx, row in enumerate(data_rows):
        r = idx + 2
        is_alt = idx % 2 == 1
        row_fill = ALT_ROW_FILL if is_alt else WHITE_FILL

        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill

        email_val = row[col_idx] if col_idx < len(row) else None
        email_str = str(email_val).strip() if email_val is not None else ""

        if not email_str:
            status = "Empty"
            empty_count += 1
            fill = GRAY_FILL
        elif _EMAIL_RE.match(email_str):
            status = "Valid"
            valid_count += 1
            fill = GREEN_FILL
        else:
            status = "Invalid"
            invalid_count += 1
            fill = RED_FILL

        sc = ws.cell(row=r, column=status_col, value=status)
        sc.fill = fill
        sc.font = WHITE_FONT
        sc.border = THIN_BORDER

    auto_size(ws)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    out.save(buf)
    base = safe_base_filename(file.filename, "workbook")

    output_name = f"email-validation-{base}.xlsx"
    resp = await record_and_respond(
        principal=principal,
        background_tasks=background_tasks,
        jobs_service=jobs_service,
        tool_slug="validate-emails",
        tool_name="Validate Emails",
        original_filename=file.filename,
        output_bytes=buf.getvalue(),
        output_filename=output_name,
        mime_type=_XLSX_MIME,
        success=True,
        error_type=None,
        duration_ms=int((time.perf_counter() - started) * 1000),
    )
    resp.headers["X-Valid-Count"] = str(valid_count)
    resp.headers["X-Invalid-Count"] = str(invalid_count)
    resp.headers["X-Empty-Count"] = str(empty_count)
    resp.headers["Access-Control-Expose-Headers"] = "X-Valid-Count, X-Invalid-Count, X-Empty-Count"
    return resp

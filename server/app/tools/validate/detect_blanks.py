from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_reader import parse_excel_bytes
from app.tools._common import check_excel_file, file_response, read_with_limit, safe_base_filename
from app.tools.analyze._styles import (
    ALT_ROW_FILL, BODY_FONT, BOLD_FONT, CENTER_ALIGNMENT, THIN_BORDER,
    WHITE_FILL, apply_header_row, auto_size, write_info_box,
)

router = APIRouter()


@router.post(
    "/detect-blanks",
    summary="Detect Blanks",
    description="Scans all sheets for blank cells and returns a report.",
)
async def detect_blanks(
    file: UploadFile = File(..., description="Excel file"),
):
    check_excel_file(file)
    raw = await read_with_limit(file)

    workbook_data = parse_excel_bytes(raw, file.filename)

    summary_rows: list[list[Any]] = []
    detail_rows: list[list[str]] = []
    total_blanks = 0
    sheets_affected = 0

    for sheet_name, rows in workbook_data.items():
        if not rows:
            continue
        header = rows[0]
        data_rows = rows[1:] if len(rows) > 1 else []
        total_cells = len(header) * len(data_rows) if data_rows else 0
        blank_count = 0

        for r_idx, row in enumerate(data_rows):
            for c_idx in range(len(header)):
                val = row[c_idx] if c_idx < len(row) else None
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    blank_count += 1
                    col_header = str(header[c_idx]) if header[c_idx] is not None else ""
                    cell_ref = f"{get_column_letter(c_idx + 1)}{r_idx + 2}"
                    detail_rows.append([sheet_name, cell_ref, str(r_idx + 2), col_header])

        pct = round(blank_count / total_cells * 100, 1) if total_cells else 0
        summary_rows.append([sheet_name, total_cells, blank_count, f"{pct}%"])
        total_blanks += blank_count
        if blank_count > 0:
            sheets_affected += 1

    out = Workbook()
    ws_sum = out.active
    ws_sum.title = "Summary"

    next_row = write_info_box(ws_sum, [
        ("File Name:", file.filename or "workbook"),
        ("Total Blanks:", str(total_blanks)),
        ("Sheets Affected:", str(sheets_affected)),
    ])

    hdr = next_row + 1
    sum_headers = ["Sheet", "Total Cells", "Blank Count", "Blank %"]
    for c, h in enumerate(sum_headers, 1):
        ws_sum.cell(row=hdr, column=c, value=h)
    apply_header_row(ws_sum, hdr, len(sum_headers))

    for idx, row_data in enumerate(summary_rows):
        r = hdr + 1 + idx
        row_fill = ALT_ROW_FILL if idx % 2 == 1 else WHITE_FILL
        for c, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill
            if c >= 2:
                cell.alignment = CENTER_ALIGNMENT

    auto_size(ws_sum)

    ws_det = out.create_sheet("Detail")
    det_headers = ["Sheet", "Cell Reference", "Row Number", "Column Header"]
    for c, h in enumerate(det_headers, 1):
        ws_det.cell(row=1, column=c, value=h)
    apply_header_row(ws_det, 1, len(det_headers))

    for idx, row_data in enumerate(detail_rows[:50000]):
        r = idx + 2
        row_fill = ALT_ROW_FILL if idx % 2 == 1 else WHITE_FILL
        for c, val in enumerate(row_data, 1):
            cell = ws_det.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill

    auto_size(ws_det)
    ws_det.freeze_panes = "A2"

    buf = BytesIO()
    out.save(buf)
    base = safe_base_filename(file.filename, "workbook")

    resp = file_response(
        buf.getvalue(),
        f"blanks-report-{base}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["X-Total-Blanks"] = str(total_blanks)
    resp.headers["X-Sheets-Affected"] = str(sheets_affected)
    resp.headers["Access-Control-Expose-Headers"] = "X-Total-Blanks, X-Sheets-Affected"
    return resp

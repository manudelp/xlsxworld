from __future__ import annotations

import time

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile
from openpyxl.utils import get_column_letter

from app.core.security import AuthenticatedPrincipal
from app.services.excel_editor import (
    load_workbook_for_edit,
    save_workbook_to_bytes,
    supports_inplace_edit,
)
from app.services.jobs_service import JobsService
from app.tools._common import check_excel_file, has_visual_elements, read_upload_for_principal
from app.tools._recording import (
    get_current_user_optional,
    jobs_service_dep,
    record_and_respond,
)

router = APIRouter()

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PADDING = 2
MIN_WIDTH = 8
MAX_WIDTH = 60


@router.post(
    "/auto-size-columns",
    summary="Auto-Size Columns",
    description="Auto-sizes all column widths based on cell content across all sheets.",
)
async def auto_size_columns(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="Excel file"),
    principal: AuthenticatedPrincipal | None = Depends(get_current_user_optional),
    jobs_service: JobsService = Depends(jobs_service_dep),
):
    started = time.perf_counter()
    check_excel_file(file)
    raw = await read_upload_for_principal(file, principal=principal)

    if not supports_inplace_edit(file.filename):
        raise HTTPException(
            status_code=400,
            detail="Auto-size columns requires an .xlsx or .xlsm file.",
        )

    loaded = load_workbook_for_edit(raw, file.filename)
    wb = loaded.workbook

    for ws in wb.worksheets:
        # Walking ws.columns on a workbook with merged cells can raise; iterate
        # over the cell ranges manually so a single problematic sheet doesn't
        # break the whole tool.
        max_col = ws.max_column
        max_row = ws.max_row
        if max_col == 0 or max_row == 0:
            continue
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            width = MIN_WIDTH
            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    width = max(width, min(len(str(value)) + PADDING, MAX_WIDTH))
            ws.column_dimensions[col_letter].width = width

    output_bytes = save_workbook_to_bytes(wb)

    return await record_and_respond(
        principal=principal,
        background_tasks=background_tasks,
        jobs_service=jobs_service,
        tool_slug="auto-size-columns",
        tool_name="Auto-Size Columns",
        original_filename=file.filename,
        output_bytes=output_bytes,
        output_filename="auto-sized.xlsx",
        mime_type=_XLSX_MIME,
        success=True,
        error_type=None,
        duration_ms=int((time.perf_counter() - started) * 1000),
        visual_elements_removed=loaded.visual_elements_lost or has_visual_elements(raw),
    )

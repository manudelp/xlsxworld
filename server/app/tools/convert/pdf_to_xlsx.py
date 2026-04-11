from __future__ import annotations

from io import BytesIO

import pdfplumber
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from openpyxl import Workbook

from app.tools._common import (
    file_response,
    read_with_limit,
    safe_base_filename,
    safe_sheet_title,
    unique_sheet_title,
)

router = APIRouter()

_PDF_CONTENT_TYPES = {"application/pdf"}


@router.post(
    "/pdf-to-xlsx",
    summary="PDF to XLSX",
    description="Extracts tables from a PDF file and converts them into an XLSX workbook.",
)
async def pdf_to_xlsx(
    file: UploadFile = File(..., description="PDF file to convert"),
    include_headers: bool = Form(True, description="Treat the first row of each table as headers"),
    one_sheet_per_page: bool = Form(False, description="Create a separate sheet for each PDF page"),
):
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()
    if not filename.endswith(".pdf") and content_type not in _PDF_CONTENT_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported file type, expected PDF")

    raw = await read_with_limit(file)

    try:
        pdf = pdfplumber.open(BytesIO(raw))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to open PDF: {exc}") from exc

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    used_names: set[str] = set()
    all_rows: list[list] = []

    for page_num, page in enumerate(pdf.pages, start=1):
        tables = page.extract_tables()
        if not tables:
            continue

        if one_sheet_per_page:
            title = unique_sheet_title(
                safe_sheet_title(f"Page {page_num}", "Page"), used_names
            )
            ws = wb.create_sheet(title=title)
            for table in tables:
                for row in table:
                    ws.append([cell.strip() if cell else "" for cell in row])
                ws.append([])  # blank row between tables on same page
        else:
            for table in tables:
                for row in table:
                    all_rows.append([cell.strip() if cell else "" for cell in row])

    if not one_sheet_per_page:
        if not all_rows:
            raise HTTPException(status_code=400, detail="No tables found in the PDF file")
        title = unique_sheet_title(safe_sheet_title("Sheet1", "Sheet"), used_names)
        ws = wb.create_sheet(title=title)
        for row in all_rows:
            ws.append(row)

    if len(wb.sheetnames) == 0:
        raise HTTPException(status_code=400, detail="No tables found in the PDF file")

    pdf.close()

    output = BytesIO()
    wb.save(output)

    out_name = f"{safe_base_filename(file.filename, 'converted')}.xlsx"

    return file_response(
        output.getvalue(),
        out_name,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

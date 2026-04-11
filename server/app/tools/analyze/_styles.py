from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
INFO_LABEL_FONT = Font(name="Arial", bold=True, size=10, color="1F3864")
INFO_VALUE_FONT = Font(name="Arial", size=10)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="center")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

GREEN_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
RED_FILL = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
AMBER_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
BLUE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TEAL_FILL = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
GRAY_FILL = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
WHITE_BOLD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)


def auto_size(ws: Worksheet, *, min_width: float = 10, max_width: float = 50) -> None:
    for col_cells in ws.columns:
        length = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)) + 3, max_width))
        ws.column_dimensions[col_letter].width = length


def apply_header_row(ws: Worksheet, row_num: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def write_info_box(ws: Worksheet, pairs: list[tuple[str, str]], start_row: int = 1) -> int:
    for i, (label, value) in enumerate(pairs):
        r = start_row + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = INFO_LABEL_FONT
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = INFO_VALUE_FONT
    return start_row + len(pairs)


def write_data_row(
    ws: Worksheet,
    row_num: int,
    values: list,
    col_count: int,
    *,
    is_alt: bool = False,
    font: Font = BODY_FONT,
) -> None:
    row_fill = ALT_ROW_FILL if is_alt else WHITE_FILL
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = font
        cell.border = THIN_BORDER
        cell.fill = row_fill
    for col in range(len(values) + 1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = THIN_BORDER
        cell.fill = row_fill


def write_reference_tab(
    ws: Worksheet,
    headers: list[str],
    rows: list[list[str]],
) -> None:
    col_count = len(headers)
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    apply_header_row(ws, 1, col_count)

    for idx, row_data in enumerate(rows):
        r = idx + 2
        is_alt = idx % 2 == 1
        row_fill = ALT_ROW_FILL if is_alt else WHITE_FILL
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = BOLD_FONT if col == 1 else BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill
            cell.alignment = WRAP_ALIGNMENT

    auto_size(ws, max_width=55)
    ws.freeze_panes = "A2"

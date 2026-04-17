from __future__ import annotations

import time
from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core.security import AuthenticatedPrincipal
from app.services.jobs_service import JobsService
from app.tools._common import (
    check_excel_file, read_with_limit,
    safe_base_filename, unique_sheet_title,
)
from app.tools._recording import (
    get_current_user_optional,
    jobs_service_dep,
    record_and_respond,
)
from app.tools.analyze._styles import (
    ALT_ROW_FILL, BODY_FONT, BOLD_FONT, CENTER_ALIGNMENT, THIN_BORDER,
    WHITE_FILL, GREEN_FILL, RED_FILL, AMBER_FILL, WHITE_BOLD_FONT,
    apply_header_row, auto_size, write_info_box, write_reference_tab,
)

router = APIRouter()

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_CHANGE_STYLES: dict[str, tuple[PatternFill, Font]] = {
    "Added": (GREEN_FILL, WHITE_BOLD_FONT),
    "Removed": (RED_FILL, WHITE_BOLD_FONT),
    "Modified": (AMBER_FILL, WHITE_BOLD_FONT),
}

_REFERENCE_ROWS: list[list[str]] = [
    ["Summary Tab", "High-level overview", "Shows total counts of sheets added, removed, modified, and total changed cells.", "Start here to understand the scope of changes."],
    ["Per-Sheet Tabs", "Cell-level diff for each modified sheet", "Each tab shows every changed cell with its original and new value/formula.", "Review these tabs to see exactly what changed in each sheet."],
    ["Added (green)", "Cell exists only in the modified file", "A new value or formula was introduced that did not exist in the original.", "Verify the addition is intentional."],
    ["Removed (red)", "Cell exists only in the original file", "A value or formula was deleted from the modified version.", "Confirm the deletion was intended and does not break dependencies."],
    ["Modified (amber)", "Cell exists in both but value or formula differs", "The cell content was changed between the two versions.", "Compare original vs. new to validate the change."],
    ["Sheets Added", "Entire sheet is new in the modified file", "A new tab was created that did not exist in the original workbook.", "Review the new sheet for correctness."],
    ["Sheets Removed", "Entire sheet is missing from the modified file", "A tab that existed in the original was deleted.", "Confirm the sheet removal was deliberate."],
]


def _read_sheet_data(ws_val, ws_form) -> dict[str, tuple[str, str]]:
    data: dict[str, tuple[str, str]] = {}
    for row_val, row_form in zip(ws_val.iter_rows(), ws_form.iter_rows()):
        for cv, cf in zip(row_val, row_form):
            if cv.value is None and (cf.value is None or cf.value == ""):
                continue
            coord = getattr(cv, "coordinate", None) or getattr(cf, "coordinate", None) or ""
            if not coord:
                continue
            val = str(cv.value) if cv.value is not None else ""
            formula = str(cf.value) if cf.value is not None else ""
            data[coord] = (val, formula)
    return data


@router.post(
    "/compare-workbooks",
    summary="Compare Workbooks",
    description="Compare two Excel files cell by cell and return an XLSX diff report.",
)
async def compare_workbooks(
    background_tasks: BackgroundTasks,
    file_a: UploadFile = File(..., description="Original Excel file"),
    file_b: UploadFile = File(..., description="Modified Excel file"),
    principal: AuthenticatedPrincipal | None = Depends(get_current_user_optional),
    jobs_service: JobsService = Depends(jobs_service_dep),
):
    started = time.perf_counter()
    check_excel_file(file_a)
    check_excel_file(file_b)
    raw_a = await read_with_limit(file_a)
    raw_b = await read_with_limit(file_b)

    try:
        wb_a_val = load_workbook(BytesIO(raw_a), data_only=True, read_only=True)
        wb_a_form = load_workbook(BytesIO(raw_a), data_only=False, read_only=True)
        wb_b_val = load_workbook(BytesIO(raw_b), data_only=True, read_only=True)
        wb_b_form = load_workbook(BytesIO(raw_b), data_only=False, read_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="Could not read one of the workbooks. The file may be corrupted or in an unsupported format.",
        ) from exc

    sheets_a = {ws.title for ws in wb_a_val.worksheets}
    sheets_b = {ws.title for ws in wb_b_val.worksheets}
    added_sheets = sorted(sheets_b - sheets_a)
    removed_sheets = sorted(sheets_a - sheets_b)
    common_sheets = sorted(sheets_a & sheets_b)

    out = Workbook()
    used_titles: set[str] = {"Summary", "How to Read This Report"}
    total_changed = 0
    modified_sheets: list[str] = []
    per_sheet_stats: list[dict] = []

    for sheet_name in common_sheets:
        data_a = _read_sheet_data(wb_a_val[sheet_name], wb_a_form[sheet_name])
        data_b = _read_sheet_data(wb_b_val[sheet_name], wb_b_form[sheet_name])
        all_cells = sorted(set(data_a.keys()) | set(data_b.keys()))
        changes: list[list[str]] = []

        for coord in all_cells:
            in_a = coord in data_a
            in_b = coord in data_b
            if in_a and in_b:
                val_a, form_a = data_a[coord]
                val_b, form_b = data_b[coord]
                if val_a != val_b or form_a != form_b:
                    changes.append([coord, "Modified", val_a, val_b, form_a, form_b])
            elif in_a:
                val_a, form_a = data_a[coord]
                changes.append([coord, "Removed", val_a, "", form_a, ""])
            else:
                val_b, form_b = data_b[coord]
                changes.append([coord, "Added", "", val_b, "", form_b])

        if not changes:
            continue

        modified_sheets.append(sheet_name)
        total_changed += len(changes)

        added_count = sum(1 for c in changes if c[1] == "Added")
        removed_count = sum(1 for c in changes if c[1] == "Removed")
        modified_count = sum(1 for c in changes if c[1] == "Modified")
        per_sheet_stats.append({
            "sheet": sheet_name,
            "added": added_count,
            "removed": removed_count,
            "modified": modified_count,
            "total": len(changes),
        })

        title = unique_sheet_title(sheet_name[:31], used_titles)
        ws_diff = out.create_sheet(title)
        diff_headers = ["Cell Reference", "Change Type", "Original Value", "New Value", "Original Formula", "New Formula"]
        for c, h in enumerate(diff_headers, 1):
            ws_diff.cell(row=1, column=c, value=h)
        apply_header_row(ws_diff, 1, len(diff_headers))

        for idx, row_data in enumerate(changes):
            r = idx + 2
            is_alt = idx % 2 == 1
            row_fill = ALT_ROW_FILL if is_alt else WHITE_FILL
            change_type = row_data[1]

            for c, val in enumerate(row_data, 1):
                cell = ws_diff.cell(row=r, column=c, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if c == 2:
                    style = _CHANGE_STYLES.get(change_type)
                    if style:
                        cell.fill = style[0]
                        cell.font = style[1]
                    cell.alignment = CENTER_ALIGNMENT
                else:
                    cell.fill = row_fill

        auto_size(ws_diff, max_width=50)
        ws_diff.freeze_panes = "A2"

    wb_a_val.close()
    wb_a_form.close()
    wb_b_val.close()
    wb_b_form.close()

    # ── Summary tab ──────────────────────────────────────────────────
    ws_sum = out.active
    ws_sum.title = "Summary"

    name_a = file_a.filename or "original"
    name_b = file_b.filename or "modified"
    scan_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    next_row = write_info_box(ws_sum, [
        ("Original File:", name_a),
        ("Modified File:", name_b),
        ("Comparison Date:", scan_time),
        ("Sheets in Original:", str(len(sheets_a))),
        ("Sheets in Modified:", str(len(sheets_b))),
    ])

    # ── High-level metrics ───────────────────────────────────────────
    metrics_hdr = next_row + 1
    m_headers = ["Metric", "Count"]
    for c, h in enumerate(m_headers, 1):
        ws_sum.cell(row=metrics_hdr, column=c, value=h)
    apply_header_row(ws_sum, metrics_hdr, len(m_headers))

    metric_rows: list[tuple[str, int, PatternFill | None, Font | None]] = [
        ("Sheets Added", len(added_sheets), GREEN_FILL if added_sheets else None, WHITE_BOLD_FONT if added_sheets else None),
        ("Sheets Removed", len(removed_sheets), RED_FILL if removed_sheets else None, WHITE_BOLD_FONT if removed_sheets else None),
        ("Sheets Modified", len(modified_sheets), AMBER_FILL if modified_sheets else None, WHITE_BOLD_FONT if modified_sheets else None),
        ("Total Changed Cells", total_changed, None, BOLD_FONT),
    ]

    r = metrics_hdr + 1
    for label, count, fill, font in metric_rows:
        lc = ws_sum.cell(row=r, column=1, value=label)
        lc.font = BODY_FONT
        lc.border = THIN_BORDER
        vc = ws_sum.cell(row=r, column=2, value=count)
        vc.font = font or BODY_FONT
        vc.border = THIN_BORDER
        vc.alignment = CENTER_ALIGNMENT
        if fill:
            vc.fill = fill
        r += 1

    # ── Added / removed sheet names ──────────────────────────────────
    if added_sheets:
        r += 1
        ws_sum.cell(row=r, column=1, value="Added Sheet Names:").font = BOLD_FONT
        ws_sum.cell(row=r, column=2, value=", ".join(added_sheets)).font = BODY_FONT
    if removed_sheets:
        r += 1
        ws_sum.cell(row=r, column=1, value="Removed Sheet Names:").font = BOLD_FONT
        ws_sum.cell(row=r, column=2, value=", ".join(removed_sheets)).font = BODY_FONT

    # ── Per-sheet breakdown ──────────────────────────────────────────
    if per_sheet_stats:
        r += 2
        ps_headers = ["Sheet", "Added", "Removed", "Modified", "Total Changes"]
        for c, h in enumerate(ps_headers, 1):
            ws_sum.cell(row=r, column=c, value=h)
        apply_header_row(ws_sum, r, len(ps_headers))
        r += 1

        for idx, ps in enumerate(per_sheet_stats):
            is_alt = idx % 2 == 1
            row_fill = ALT_ROW_FILL if is_alt else WHITE_FILL
            vals = [ps["sheet"], ps["added"], ps["removed"], ps["modified"], ps["total"]]
            for c, val in enumerate(vals, 1):
                cell = ws_sum.cell(row=r, column=c, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.fill = row_fill
                if c >= 2:
                    cell.alignment = CENTER_ALIGNMENT
            r += 1

        # Totals
        ws_sum.cell(row=r, column=1, value="TOTAL").font = BOLD_FONT
        ws_sum.cell(row=r, column=1).border = THIN_BORDER
        for c, key in enumerate(["added", "removed", "modified", "total"], 2):
            tc = ws_sum.cell(row=r, column=c, value=sum(ps[key] for ps in per_sheet_stats))
            tc.font = BOLD_FONT
            tc.border = THIN_BORDER
            tc.alignment = CENTER_ALIGNMENT

    auto_size(ws_sum)
    ws_sum.freeze_panes = f"A{metrics_hdr + 1}"

    # Move Summary to first position
    out.move_sheet("Summary", offset=-len(out.sheetnames) + 1)

    # ── Reference tab ────────────────────────────────────────────────
    write_reference_tab(
        out.create_sheet("How to Read This Report"),
        ["Term", "What It Means", "Details", "What to Do"],
        _REFERENCE_ROWS,
    )

    # ── Save ─────────────────────────────────────────────────────────
    buf = BytesIO()
    out.save(buf)

    base_a = safe_base_filename(name_a, "original")
    base_b = safe_base_filename(name_b, "modified")
    output_name = f"comparison-{base_a}-vs-{base_b}.xlsx"

    resp = await record_and_respond(
        principal=principal,
        background_tasks=background_tasks,
        jobs_service=jobs_service,
        tool_slug="compare-workbooks",
        tool_name="Compare Workbooks",
        original_filename=file_a.filename,
        output_bytes=buf.getvalue(),
        output_filename=output_name,
        mime_type=_XLSX_MIME,
        success=True,
        error_type=None,
        duration_ms=int((time.perf_counter() - started) * 1000),
    )
    resp.headers["X-Sheets-Added"] = str(len(added_sheets))
    resp.headers["X-Sheets-Removed"] = str(len(removed_sheets))
    resp.headers["X-Sheets-Modified"] = str(len(modified_sheets))
    resp.headers["X-Total-Changed"] = str(total_changed)
    resp.headers["Access-Control-Expose-Headers"] = "X-Sheets-Added, X-Sheets-Removed, X-Sheets-Modified, X-Total-Changed"
    return resp

from __future__ import annotations

import statistics
import time
from io import BytesIO
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile
from openpyxl import Workbook

from app.core.security import AuthenticatedPrincipal
from app.services.excel_reader import parse_excel_bytes
from app.services.jobs_service import JobsService
from app.tools._common import check_excel_file, read_with_limit, safe_base_filename
from app.tools._recording import (
    get_current_user_optional,
    jobs_service_dep,
    record_and_respond,
)
from app.tools.analyze._styles import (
    ALT_ROW_FILL, BODY_FONT, BOLD_FONT, CENTER_ALIGNMENT, THIN_BORDER,
    WHITE_FILL, apply_header_row, auto_size,
)

router = APIRouter()

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _numeric_values(cells: list[Any]) -> list[float]:
    nums: list[float] = []
    for c in cells:
        if c is None:
            continue
        if isinstance(c, (int, float)):
            nums.append(float(c))
        elif isinstance(c, str):
            try:
                nums.append(float(c))
            except ValueError:
                pass
    return nums


@router.post(
    "/summary-stats",
    summary="Summary Stats",
    description="Compute summary statistics for numeric columns across all sheets.",
)
async def summary_stats(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="Excel file"),
    principal: AuthenticatedPrincipal | None = Depends(get_current_user_optional),
    jobs_service: JobsService = Depends(jobs_service_dep),
):
    started = time.perf_counter()
    check_excel_file(file)
    raw = await read_with_limit(file)

    workbook_data = parse_excel_bytes(raw, file.filename)

    out = Workbook()
    out.remove(out.active)
    total_columns = 0
    sheets_analyzed = 0

    for sheet_name, rows in workbook_data.items():
        if len(rows) < 2:
            continue
        header = rows[0]
        data_rows = rows[1:]
        stats_rows: list[list[Any]] = []

        for col_idx, col_name in enumerate(header):
            col_cells = [r[col_idx] if col_idx < len(r) else None for r in data_rows]
            nums = _numeric_values(col_cells)
            if not nums:
                continue
            total_columns += 1
            blanks = sum(1 for c in col_cells if c is None or (isinstance(c, str) and c.strip() == ""))
            stats_rows.append([
                str(col_name) if col_name is not None else f"Column {col_idx + 1}",
                len(nums),
                round(min(nums), 4),
                round(max(nums), 4),
                round(statistics.mean(nums), 4),
                round(statistics.median(nums), 4),
                round(statistics.stdev(nums), 4) if len(nums) > 1 else 0,
                round(sum(nums), 4),
                blanks,
            ])

        if not stats_rows:
            continue
        sheets_analyzed += 1
        ws = out.create_sheet(sheet_name[:31])
        headers = ["Column", "Count", "Min", "Max", "Mean", "Median", "Std Dev", "Sum", "Blanks"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        apply_header_row(ws, 1, len(headers))

        for idx, row_data in enumerate(stats_rows):
            r = idx + 2
            row_fill = ALT_ROW_FILL if idx % 2 == 1 else WHITE_FILL
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.fill = row_fill
                if c >= 2:
                    cell.alignment = CENTER_ALIGNMENT

        auto_size(ws)
        ws.freeze_panes = "A2"

    if not out.worksheets:
        ws = out.create_sheet("No Data")
        ws.cell(row=1, column=1, value="No numeric columns found.")

    buf = BytesIO()
    out.save(buf)
    base = safe_base_filename(file.filename, "workbook")

    output_name = f"summary-stats-{base}.xlsx"
    resp = await record_and_respond(
        principal=principal,
        background_tasks=background_tasks,
        jobs_service=jobs_service,
        tool_slug="summary-stats",
        tool_name="Summary Stats",
        original_filename=file.filename,
        output_bytes=buf.getvalue(),
        output_filename=output_name,
        mime_type=_XLSX_MIME,
        success=True,
        error_type=None,
        duration_ms=int((time.perf_counter() - started) * 1000),
    )
    resp.headers["X-Sheets-Analyzed"] = str(sheets_analyzed)
    resp.headers["X-Columns-Found"] = str(total_columns)
    resp.headers["Access-Control-Expose-Headers"] = "X-Sheets-Analyzed, X-Columns-Found"
    return resp

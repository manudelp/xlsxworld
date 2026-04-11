from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.tools._common import check_excel_file, file_response, read_with_limit, safe_base_filename
from app.tools.analyze._styles import (
    ALT_ROW_FILL, BODY_FONT, BOLD_FONT, CENTER_ALIGNMENT, THIN_BORDER,
    WHITE_FILL, WRAP_ALIGNMENT,
    apply_header_row, auto_size, write_info_box, write_reference_tab,
)

router = APIRouter()

FORMULA_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}

_SEVERITY: dict[str, str] = {
    "#REF!": "High", "#NAME?": "High",
    "#VALUE!": "Medium", "#NUM!": "Medium",
    "#DIV/0!": "Low", "#N/A": "Low", "#NULL!": "Low",
}

_FIX_HINTS: dict[str, str] = {
    "#DIV/0!": "Denominator is zero or empty. Wrap with =IFERROR() or check divisor.",
    "#REF!": "Cell reference is invalid. A referenced cell may have been deleted or moved.",
    "#VALUE!": "Wrong data type in formula. Check for text in numeric cells.",
    "#NAME?": "Formula name not recognized. Check for typos in function names.",
    "#N/A": "Value not found. Check VLOOKUP/MATCH source range.",
    "#NUM!": "Invalid numeric value. Check for impossible calculations (e.g. SQRT of negative).",
    "#NULL!": "Incorrect range operator. Check for missing colon or comma in range.",
}

_ERROR_TYPE_FILLS: dict[str, PatternFill] = {
    "#REF!": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "#DIV/0!": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
    "#VALUE!": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
    "#NAME?": PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid"),
    "#N/A": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
    "#NUM!": PatternFill(start_color="008080", end_color="008080", fill_type="solid"),
    "#NULL!": PatternFill(start_color="808080", end_color="808080", fill_type="solid"),
}

_SEVERITY_STYLES: dict[str, tuple[PatternFill, Font]] = {
    "High": (
        PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        Font(name="Arial", bold=True, color="FFFFFF"),
    ),
    "Medium": (
        PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        Font(name="Arial", bold=True, color="FFFFFF"),
    ),
    "Low": (
        PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        Font(name="Arial", bold=True, color="000000"),
    ),
}

_REFERENCE_ROWS: list[list[str]] = [
    ["#DIV/0!", "Division by zero", "A formula divides a value by zero or by an empty cell.", "Add an IFERROR wrapper or an IF check to handle zero/empty divisors before dividing."],
    ["#REF!", "Invalid cell reference", "A formula refers to a cell that has been deleted, or a paste operation shifted references out of range.", "Undo the deletion, or manually correct the formula to point to the right cell."],
    ["#VALUE!", "Wrong value type", "A formula expects a number but receives text, or an argument type is incompatible.", "Ensure all inputs are the correct type. Use VALUE(), TRIM(), or CLEAN() to convert text."],
    ["#NAME?", "Unrecognized name", "A function name is misspelled, or a named range does not exist in this workbook.", "Check spelling of function names and verify that all named ranges are defined."],
    ["#N/A", "Value not available", "A lookup function (VLOOKUP, MATCH, INDEX) cannot find the requested value.", "Verify the lookup value exists in the source range. Use IFERROR or IFNA to handle misses."],
    ["#NUM!", "Invalid number", "A formula produces a number that is too large, too small, or mathematically impossible (e.g. SQRT of a negative).", "Check input values for validity. Add guards for edge cases like negatives or overflow."],
    ["#NULL!", "Null intersection", "Two ranges that should intersect do not, usually because of a missing colon or comma.", "Replace the space between ranges with a colon (:) for a range or a comma (,) for a union."],
]


@router.post(
    "/scan-formula-errors",
    summary="Scan Formula Errors",
    description="Scan every cell across all sheets for formula errors and return an XLSX report.",
)
async def scan_formula_errors(
    file: UploadFile = File(..., description="Excel file"),
):
    check_excel_file(file)
    raw = await read_with_limit(file)

    try:
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        wb_formulas = load_workbook(BytesIO(raw), data_only=False, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {exc}") from exc

    details: list[dict[str, str]] = []
    error_counts: dict[str, int] = {}
    error_sheets: dict[str, set[str]] = {}
    total_sheets = len(wb.worksheets)

    for ws_val, ws_form in zip(wb.worksheets, wb_formulas.worksheets):
        sheet_name = ws_val.title
        for row_val, row_form in zip(ws_val.iter_rows(), ws_form.iter_rows()):
            for cell_val, cell_form in zip(row_val, row_form):
                val = cell_val.value
                if val is None:
                    continue
                val_str = str(val)
                if val_str not in FORMULA_ERRORS:
                    continue
                formula = str(cell_form.value) if cell_form.value is not None else ""
                details.append({
                    "sheet": sheet_name,
                    "cell": cell_val.coordinate or "",
                    "error_type": val_str,
                    "formula": formula,
                })
                error_counts[val_str] = error_counts.get(val_str, 0) + 1
                error_sheets.setdefault(val_str, set()).add(sheet_name)

    wb.close()
    wb_formulas.close()

    total_errors = len(details)
    scan_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    original_name = file.filename or "workbook"
    out = Workbook()

    # ── Summary ──────────────────────────────────────────────────────
    ws_sum = out.active
    ws_sum.title = "Summary"
    next_row = write_info_box(ws_sum, [
        ("File Name:", original_name),
        ("Scan Date:", scan_time),
        ("Sheets Scanned:", str(total_sheets)),
        ("Total Errors Found:", str(total_errors)),
    ])

    hdr_row = next_row + 1
    headers = ["Error Type", "Count", "% of Total Errors", "Affected Sheets", "Severity"]
    for c, h in enumerate(headers, 1):
        ws_sum.cell(row=hdr_row, column=c, value=h)
    apply_header_row(ws_sum, hdr_row, len(headers))

    r = hdr_row + 1
    for err_type, count in sorted(error_counts.items(), key=lambda x: -x[1]):
        pct = round(count / total_errors * 100, 1) if total_errors else 0
        severity = _SEVERITY.get(err_type, "Low")

        ws_sum.cell(row=r, column=1, value=err_type).font = BODY_FONT
        ws_sum.cell(row=r, column=2, value=count).font = BODY_FONT
        pc = ws_sum.cell(row=r, column=3, value=f"{pct}%")
        pc.font = BODY_FONT
        pc.alignment = CENTER_ALIGNMENT
        ws_sum.cell(row=r, column=4, value=", ".join(sorted(error_sheets.get(err_type, set())))).font = BODY_FONT
        sc = ws_sum.cell(row=r, column=5, value=severity)
        sf, sfn = _SEVERITY_STYLES.get(severity, (WHITE_FILL, BODY_FONT))
        sc.fill = sf
        sc.font = sfn
        sc.alignment = CENTER_ALIGNMENT
        for c in range(1, len(headers) + 1):
            ws_sum.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    ws_sum.cell(row=r, column=1, value="TOTAL").font = BOLD_FONT
    ws_sum.cell(row=r, column=2, value=total_errors).font = BOLD_FONT
    tc = ws_sum.cell(row=r, column=3, value="100%")
    tc.font = BOLD_FONT
    tc.alignment = CENTER_ALIGNMENT
    for c in range(1, len(headers) + 1):
        ws_sum.cell(row=r, column=c).border = THIN_BORDER

    auto_size(ws_sum)
    ws_sum.freeze_panes = f"A{hdr_row + 1}"

    # ── Detail ───────────────────────────────────────────────────────
    ws_det = out.create_sheet("Detail")
    det_headers = ["Sheet", "Cell Reference", "Error Type", "Formula", "Suggested Fix Hint"]
    for c, h in enumerate(det_headers, 1):
        ws_det.cell(row=1, column=c, value=h)
    apply_header_row(ws_det, 1, len(det_headers))

    sorted_details = sorted(details, key=lambda d: (d["sheet"], d["error_type"], d["cell"]))
    for idx, d in enumerate(sorted_details):
        r = idx + 2
        is_alt = idx % 2 == 1
        row_fill = ALT_ROW_FILL if is_alt else WHITE_FILL

        ws_det.cell(row=r, column=1, value=d["sheet"]).font = BODY_FONT
        ws_det.cell(row=r, column=2, value=d["cell"]).font = BODY_FONT
        et = ws_det.cell(row=r, column=3, value=d["error_type"])
        ef = _ERROR_TYPE_FILLS.get(d["error_type"])
        if ef:
            et.fill = ef
            et.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        else:
            et.font = BODY_FONT
        et.alignment = CENTER_ALIGNMENT
        ws_det.cell(row=r, column=4, value=d["formula"]).font = BODY_FONT
        hc = ws_det.cell(row=r, column=5, value=_FIX_HINTS.get(d["error_type"], ""))
        hc.font = BODY_FONT
        hc.alignment = WRAP_ALIGNMENT
        for c in range(1, len(det_headers) + 1):
            cell = ws_det.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if c != 3:
                cell.fill = row_fill

    auto_size(ws_det, max_width=60)
    ws_det.freeze_panes = "A2"

    # ── Reference ────────────────────────────────────────────────────
    write_reference_tab(
        out.create_sheet("How to Read This Report"),
        ["Error Type", "Meaning", "Common Cause", "How to Fix"],
        _REFERENCE_ROWS,
    )

    # ── Save ─────────────────────────────────────────────────────────
    buf = BytesIO()
    out.save(buf)
    base = safe_base_filename(original_name, "workbook")

    resp = file_response(
        buf.getvalue(),
        f"formula-errors-{base}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["X-Total-Errors"] = str(total_errors)
    resp.headers["X-Error-Breakdown"] = json.dumps(error_counts)
    resp.headers["Access-Control-Expose-Headers"] = "X-Total-Errors, X-Error-Breakdown"
    return resp

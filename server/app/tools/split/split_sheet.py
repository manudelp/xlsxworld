from __future__ import annotations

import re
from io import BytesIO

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from openpyxl import Workbook

from app.services.excel_reader import parse_excel_bytes
from app.tools._common import check_excel_file, file_response, read_with_limit, safe_sheet_title, unique_sheet_title

router = APIRouter()


def _alpha_token(index: int, lowercase: bool = False) -> str:
    chars: list[str] = []
    n = max(1, index)
    while n > 0:
        n -= 1
        chars.append(chr(ord("A") + (n % 26)))
        n //= 26
    token = "".join(reversed(chars))
    return token.lower() if lowercase else token


def _roman_token(index: int, lowercase: bool = False) -> str:
    n = max(1, index)
    numerals = [
        (1000, "M"), (900, "CM"), (500, "D"), (400, "CD"), (100, "C"),
        (90, "XC"), (50, "L"), (40, "XL"), (10, "X"), (9, "IX"),
        (5, "V"), (4, "IV"), (1, "I"),
    ]
    parts: list[str] = []
    for value, symbol in numerals:
        while n >= value:
            parts.append(symbol)
            n -= value
    token = "".join(parts)
    return token.lower() if lowercase else token


def _build_part_token(index: int, numbering_style: str, custom_tokens: list[str]) -> str:
    if numbering_style == "numeric":
        return str(index)
    if numbering_style == "numeric-padded":
        return f"{index:02d}"
    if numbering_style == "alpha-upper":
        return _alpha_token(index, lowercase=False)
    if numbering_style == "alpha-lower":
        return _alpha_token(index, lowercase=True)
    if numbering_style == "roman-upper":
        return _roman_token(index, lowercase=False)
    if numbering_style == "roman-lower":
        return _roman_token(index, lowercase=True)
    if numbering_style == "custom":
        if index <= len(custom_tokens):
            return custom_tokens[index - 1]
        return str(index)
    return str(index)


_VALID_STYLES = {
    "numeric", "numeric-padded", "alpha-upper", "alpha-lower",
    "roman-upper", "roman-lower", "custom",
}


@router.post(
    "/split-sheet",
    summary="Split Sheet",
    description="Splits one sheet into multiple sheets by row chunk size and naming strategy.",
)
async def split_sheet(
    file: UploadFile = File(..., description="Excel file"),
    sheet: str = Form(..., description="Sheet name"),
    chunk_size: int = Form(1000, description="Max rows per chunk (including header)"),
    part_base: str = Form("part", description="Base name used for split sheets"),
    part_separator: str = Form("_", description="Separator between base name and part token"),
    numbering_style: str = Form("numeric", description="Part token style"),
    custom_sequence: str = Form("", description="Custom tokens (one per line) when style is custom"),
):
    check_excel_file(file)
    raw = await read_with_limit(file)
    if chunk_size < 2:
        raise HTTPException(status_code=400, detail="chunk_size must be >= 2")

    workbook_data = parse_excel_bytes(raw, file.filename)

    if sheet not in workbook_data:
        raise HTTPException(status_code=404, detail="Sheet not found")

    rows = workbook_data[sheet]
    if not rows:
        raise HTTPException(status_code=400, detail="Sheet is empty")

    if numbering_style not in _VALID_STYLES:
        raise HTTPException(status_code=400, detail="Invalid numbering_style")

    if len(part_separator) > 4:
        raise HTTPException(status_code=400, detail="part_separator must be 4 characters or fewer")

    custom_tokens = [
        token.strip()
        for token in re.split(r"[\r\n,]+", custom_sequence)
        if token.strip()
    ]
    if numbering_style == "custom" and not custom_tokens:
        raise HTTPException(status_code=400, detail="custom_sequence is required when numbering_style is custom")

    header = rows[0]
    chunks = [rows[i : i + chunk_size] for i in range(1, len(rows), chunk_size)]

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    used_titles: set[str] = set()

    for idx, chunk in enumerate(chunks, start=1):
        token = _build_part_token(idx, numbering_style, custom_tokens)
        default_name = f"part_{idx}"
        requested_name = f"{part_base}{part_separator}{token}"
        base_title = safe_sheet_title(requested_name, default_name)
        sheet_title = unique_sheet_title(base_title, used_titles)
        part = out_wb.create_sheet(sheet_title)
        part.append(["" if v is None else v for v in header])
        for row in chunk:
            part.append(["" if v is None else v for v in row])

    if not chunks:
        token = _build_part_token(1, numbering_style, custom_tokens)
        requested_name = f"{part_base}{part_separator}{token}"
        base_title = safe_sheet_title(requested_name, "part_1")
        sheet_title = unique_sheet_title(base_title, used_titles)
        part = out_wb.create_sheet(sheet_title)
        part.append(["" if v is None else v for v in header])

    output = BytesIO()
    out_wb.save(output)

    return file_response(
        output.getvalue(),
        "split.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
